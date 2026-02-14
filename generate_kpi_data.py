"""
Generate dummy KPI dataset with Exact Formula, Required Data, Columns Used, and Formula/Logic.
Exports to xlsx with auto-fit columns for best fit.
"""
import random
import re
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    pd = None

random.seed(42)

# Normalize KPI name for matching (lowercase, single spaces, no punctuation)
def norm(s):
    return re.sub(r"\s+", " ", re.sub(r"[;/]", " ", str(s).lower())).strip()

# Formula, Required Data, Columns Used, Formula/Logic — key = normalized KPI name (substring match)
FORMULA_MAP = {
    "overloaded dts identified": (
        "% Loading = (√3 × Vavg × Iavg /1000 ÷ Rated kVA) ×100 → Flag if >80%",
        "DT Block: L1/L2/L3 Voltage, L1/L2/L3 Current; Asset Master: Rated Capacity (kVA)",
        "L1/L2/L3 Voltage, L1/L2/L3 Current, Rated kVA",
        "Flag when % Loading > 80%",
    ),
    "transformer utilization rate": (
        "(Actual kVA Load ÷ Rated Capacity) ×100",
        "DT Block: Voltage & Current; Asset Master: Rated kVA",
        "Voltage, Current, Rated kVA",
        "Actual kVA Load ÷ Rated Capacity × 100",
    ),
    "reduction in theft events": (
        "((Baseline Theft − Current Theft) ÷ Baseline Theft) ×100",
        "Event Log: Theft/Tamper confirmed events",
        "Event type, Timestamp",
        "Baseline vs Current period theft events",
    ),
    "tamper sequence detection": (
        "Pattern detection within time window (e.g., Power Fail → Cover Open → Reverse Current)",
        "Event Profile (IS 15959): Meter ID, Timestamp, Event Code, Event Status",
        "Meter ID, Timestamp, Event Code, Event Status",
        "Sequence pattern match in time window",
    ),
    "power factor deterioration": (
        "PF = kWh ÷ kVAh; Flag if PF <0.85 sustained",
        "Block Profile: kWh(Imp), kVAh(Imp)",
        "kWh(Imp), kVAh(Imp)",
        "PF = kWh / kVAh; flag when PF < 0.85 sustained",
    ),
    "overload md breach risk": (
        "(Max Demand ÷ Sanctioned Load) ×100; Flag if >90%",
        "Consumer Profile: Max Demand; Consumer Master: Sanctioned Load",
        "Max Demand, Sanctioned Load",
        "Flag when (Max Demand / Sanctioned Load) × 100 > 90%",
    ),
    "real-time phase load monitoring": (
        "Phase kVA = Vphase × Iphase ÷1000",
        "DT Block: L1/L2/L3 Voltage & Current",
        "L1/L2/L3 Voltage, L1/L2/L3 Current",
        "Per-phase kVA = V × I / 1000",
    ),
    "phase transfer recommendation": (
        "Imbalance % = (Max Phase − Avg Phase) ÷ Avg Phase ×100",
        "DT Block: Phase currents; Consumer Phase Mapping",
        "L1/L2/L3 Current, Consumer-Phase mapping",
        "Recommend phase transfer when imbalance % exceeds threshold",
    ),
    "lt loss": (
        "(DT Energy − Σ Consumer Energy) ÷ DT Energy ×100",
        "DT Block Energy; Consumer Block Energy",
        "DT Energy, Consumer Energy (sum)",
        "Energy loss % at DT",
    ),
    "billing efficiency": (
        "(Billed Energy ÷ Total Energy Supplied) ×100",
        "Billing DB: Billed Units; Energy Input Data",
        "Billed Units, Total Energy Supplied",
        "Billed ÷ Supplied × 100",
    ),
    "collection efficiency": (
        "(Amount Collected ÷ Amount Billed) ×100",
        "Billing System Data",
        "Amount Collected, Amount Billed",
        "Collected ÷ Billed × 100",
    ),
    "at&c loss": (
        "AT&C = 1 − (Billing Eff × Collection Eff)",
        "Billing + Collection Data",
        "Billing Efficiency, Collection Efficiency",
        "1 − (Billing Eff × Collection Eff)",
    ),
    "revenue recovery improvement": (
        "((Recovered ₹ − Baseline ₹) ÷ Baseline ₹) ×100",
        "Revenue DB",
        "Recovered Amount, Baseline Amount",
        "Recovered vs Baseline revenue %",
    ),
    "saidi": (
        "Total Interruption Minutes ÷ Total Customers",
        "Outage Event DB; Consumer Count",
        "Interruption minutes, Total Customers",
        "Sum(Outage Minutes) / Total Customers",
    ),
    "saifi": (
        "Total Interruptions ÷ Total Customers",
        "Outage Event DB",
        "Interruption count, Total Customers",
        "Total Interruptions / Total Customers",
    ),
    "caidi": (
        "SAIDI ÷ SAIFI",
        "SAIDI, SAIFI",
        "SAIDI, SAIFI",
        "SAIDI / SAIFI",
    ),
    "caifi": (
        "Total Interruptions ÷ Affected Customers",
        "Outage DB",
        "Interruptions, Affected Customers",
        "Interruptions / Affected Customers",
    ),
    "maifi": (
        "Momentary Interruptions ÷ Total Customers",
        "Outage DB (<5 min events)",
        "Momentary interruptions, Total Customers",
        "Momentary (<5 min) / Total Customers",
    ),
    "number of outages": (
        "Count (outage events)",
        "Outage Event DB",
        "Outage Event ID, Timestamp",
        "COUNT(outage events)",
    ),
    "duration of outages": (
        "Sum (outage duration minutes)",
        "Outage Event DB",
        "Outage start, Outage end",
        "SUM(duration minutes)",
    ),
    "dt feeder reliability trends": (
        "Monthly Trend (SAIDI/SAIFI/Outage Minutes)",
        "Outage DB + Asset Mapping",
        "SAIDI, SAIFI, Outage Minutes by month",
        "Trend of SAIDI, SAIFI, outage minutes",
    ),
    "frequency deviation index": (
        "Σ |f − fnom|",
        "Frequency samples",
        "Frequency (Hz), Nominal frequency",
        "Sum of |f - fnom| over period",
    ),
    "low power factor by dt feeder": (
        "(Blocks with PF < threshold ÷ Total Blocks) ×100",
        "Block Profile: kWh, kVAh",
        "kWh, kVAh per block",
        "PF = kWh/kVAh; % blocks with PF < threshold (e.g. 0.85)",
    ),
    "dts with high failure rate": (
        "(Failures ÷ Total DTs) ×100",
        "Maintenance DB",
        "Failure count, Total DTs",
        "Failures / Total DTs × 100",
    ),
    "field inspection hit-rate": (
        "(Confirmed Cases ÷ Total Inspections) ×100",
        "Inspection Workflow System",
        "Confirmed cases, Total inspections",
        "Confirmed / Total Inspections × 100",
    ),
    "mttr": (
        "Restore Time − Fault Detection Time",
        "Event DB + Ticket System",
        "Restore Time, Fault Detection Time",
        "Mean Time To Restore",
    ),
    "alert response time": (
        "Acknowledged Time − Alert Created Time",
        "Alert System Logs",
        "Acknowledged Time, Alert Created Time",
        "Time to acknowledge alert",
    ),
    "planned outage suppression rate": (
        "(Muted Alerts ÷ Total Alerts) ×100",
        "Alert System + Outage Schedule",
        "Muted Alerts, Total Alerts",
        "Muted / Total × 100",
    ),
    "feeders with maximum outages": (
        "Rank by outage count or duration",
        "Outage DB",
        "Feeder ID, Outage count/duration",
        "Rank feeders by outage count or duration",
    ),
    "reliability improvement trend": (
        "Monthly Trend (SAIDI/SAIFI)",
        "Outage DB",
        "SAIDI, SAIFI by month",
        "Trend of SAIDI and SAIFI",
    ),
    "consumer service reliability score": (
        "Weighted Composite (SAIDI, SAIFI, Complaints)",
        "Outage DB + Complaint DB",
        "SAIDI, SAIFI, Complaint count",
        "Weighted score from SAIDI, SAIFI, complaints",
    ),
    "composite reliability score": (
        "Weighted Reliability KPIs",
        "Reliability Data",
        "SAIDI, SAIFI, CAIDI, etc.",
        "Weighted combination of reliability KPIs",
    ),
    "composite efficiency score": (
        "Weighted (Loss + Billing + Collection)",
        "Energy + Billing Data",
        "Loss %, Billing Eff, Collection Eff",
        "Weighted loss, billing, collection",
    ),
    "loading bands": (
        "Categorize % Loading into thresholds",
        "DT Block + Rated Capacity",
        "DT Load, Rated kVA",
        "Band by % Loading (e.g. 0–50, 50–80, 80–100, >100)",
    ),
    "consumers exceeding sanctioned load": (
        "Max Demand > Sanctioned Load",
        "Consumer Profile + Master Data",
        "Max Demand, Sanctioned Load",
        "Count where Max Demand > Sanctioned Load",
    ),
    "consumers with load violation": (
        "(Violators ÷ Total Consumers) ×100",
        "Consumer Demand Data",
        "Violators count, Total Consumers",
        "Violators / Total × 100",
    ),
    "communication retry counts": (
        "Count (retry events per meter)",
        "Communication Log",
        "Meter ID, Retry events",
        "COUNT(retry events) per meter",
    ),
    "dt failure rate": (
        "(Failures ÷ Total DTs) ×100",
        "Maintenance DB",
        "Failures, Total DTs",
        "Failures / Total DTs × 100",
    ),
    "top overloaded assets": (
        "Rank by % Loading or Peak Load",
        "DT Block + Rated Capacity",
        "DT Load, Rated kVA",
        "Rank by % Loading or peak kVA",
    ),
    "total assets tracked": (
        "Meters: COUNT(DISTINCT newMeterNumber); Feeders: COUNT(DISTINCT FeederCode); DTs: COUNT(DISTINCT DTRCode)",
        "Asset Master / Mapping tables",
        "newMeterNumber, FeederCode, DTRCode",
        "Meters: COUNT(DISTINCT newMeterNumber); Feeders: COUNT(DISTINCT FeederCode); DTs: COUNT(DISTINCT DTRCode)",
    ),
    "verification pending count": (
        "COUNT(*) WHERE QC status not Approved",
        "QC workflow tables",
        "QC1Status, QC2Status, QC3Status",
        "COUNT(*) WHERE QC1Status != 'Approved' OR QC2Status != 'Approved' OR QC3Status != 'Approved'",
    ),
    "correction cycle time": (
        "For each meter: Cycle Time = Final_QC_Date − installationDate; Avg = MEAN(Cycle Time)",
        "QC and installation dates",
        "QC1DoneDate, QC2DoneDate, QC3DoneDate, installationDate",
        "Cycle Time = Final_QC_Date − installationDate; Avg Cycle Time = MEAN(Cycle Time)",
    ),
    "signal strength statistics": (
        "Mean: AVG(TSP1, TSP2); Min/Max: MIN, MAX; Std Dev: STD(TSP1, TSP2)",
        "Meter communication profile",
        "TSP1 (dBm), TSP2 (dBm)",
        "Mean Signal Strength: AVG(TSP1, TSP2); Min/Max: MIN, MAX; Std Dev: STD(TSP1, TSP2)",
    ),
    "weak signal percentage": (
        "Weak % = (Count of meters where TSP < threshold / Total meters) × 100",
        "TSP1, TSP2 (dBm); threshold e.g. -90 dBm",
        "TSP1 (dBm), TSP2 (dBm)",
        "Define weak threshold (e.g. -90 dBm); Weak % = (Meters with TSP < threshold / Total meters) × 100",
    ),
    "non-reporting meters": (
        "Meter is non-reporting if: Current_Time − Last_Report_Time > 24 hours",
        "RTC, Meter ID",
        "RTC, newMeterNumber",
        "Last_Report_Time = MAX(RTC) per meter; flag if Current_Time − Last_Report_Time > 24 hours",
    ),
}

def find_formula(kpi_name):
    n = norm(kpi_name)
    for key, val in FORMULA_MAP.items():
        if key in n or n in key:
            return val
    return ("", "", "", "")

# KPI_SPECS: (Dashboard, Department, KPI Name, value_type, min_val, max_val, unit)
KPI_SPECS = [
    ("Dashboard-1", "Finance", "Feeder Loss (%)", "pct", 3, 18, "%"),
    ("Dashboard-1", "Finance", "DT (Distribution Transformer) Loss (%)", "pct", 2, 12, "%"),
    ("Dashboard-1", "Finance", "LT Loss (%)", "pct", 1, 8, "%"),
    ("Dashboard-1", "Finance", "Billing Efficiency (%)", "pct", 78, 98, "%"),
    ("Dashboard-1", "Finance", "Collection Efficiency (%)", "pct", 72, 96, "%"),
    ("Dashboard-1", "Finance", "AT&C Loss (%)", "pct", 8, 28, "%"),
    ("Dashboard-1", "Finance", "Top X Best/Worst Feeders/DTs", "count", 5, 20, "count"),
    ("Dashboard-1", "Finance", "Top High Loss DTs / Feeders", "count", 8, 35, "count"),
    ("Dashboard-1", "Finance", "Top High-Loss Feeders / DTs", "count", 6, 28, "count"),
    ("Dashboard-2", "Operation", "SAIDI", "minutes", 45, 380, "min"),
    ("Dashboard-2", "Operation", "SAIFI", "index", 2, 25, "interruptions"),
    ("Dashboard-2", "Operation", "CAIDI", "minutes", 25, 95, "min"),
    ("Dashboard-2", "Operation", "CAIFI", "index", 1, 18, "interruptions"),
    ("Dashboard-2", "Operation", "MAIFI", "index", 0.2, 8, "interruptions"),
    ("Dashboard-2", "Operation", "Number of Outages (Frequency)", "count", 12, 450, "count"),
    ("Dashboard-2", "Operation", "Duration of Outages (Minutes)", "minutes", 120, 7200, "min"),
    ("Dashboard-2", "Operation", "DT/Feeder Reliability Trends (Monthly/Yearly)", "pct", 85, 99.5, "%"),
    ("Dashboard-2", "Operation", "DTs with High Failure Rate", "count", 3, 45, "count"),
    ("Dashboard-2", "Operation", "Detection Accuracy", "pct", 82, 98, "%"),
    ("Dashboard-2", "Operation", "False Positive Rate", "pct", 1, 15, "%"),
    ("Dashboard-2", "Operation", "Field inspection hit-rate", "pct", 65, 92, "%"),
    ("Dashboard-2", "Operation", "MTTI", "minutes", 8, 95, "min"),
    ("Dashboard-2", "Operation", "MTTR", "minutes", 25, 180, "min"),
    ("Dashboard-2", "Operation", "Alert response time", "minutes", 5, 45, "min"),
    ("Dashboard-2", "Operation", "Planned outage suppression rate", "pct", 70, 98, "%"),
    ("Dashboard-2", "Analytics", "Low-voltage pockets", "count", 2, 28, "count"),
    ("Dashboard-2", "Operation", "Feeders with Maximum Outages", "count", 4, 22, "count"),
    ("Dashboard-2", "Operation", "Reliability Improvement Trend", "pct", 2, 18, "%"),
    ("Dashboard-2", "Operation", "Consumer Service Reliability Score", "score", 72, 95, "score"),
    ("Dashboard-2", "Operation", "Composite Reliability Score", "score", 68, 94, "score"),
    ("Dashboard-2", "Operation", "Composite Efficiency Score", "score", 65, 92, "score"),
    ("Dashboard-3", "Technical", "% DT Peak Loading", "pct", 45, 98, "%"),
    ("Dashboard-3", "Technical", "% DT Loading", "pct", 38, 92, "%"),
    ("Dashboard-3", "Technical", "DT Load (kVA)", "kva", 25, 315, "kVA"),
    ("Dashboard-3", "Technical", "% Loading Bands", "pct", 0, 100, "%"),
    ("Dashboard-3", "Technical", "Top Overloaded DTs / Feeders", "count", 5, 30, "count"),
    ("Dashboard-3", "Technical", "Load Rise Trend", "pct", 2, 22, "%"),
    ("Dashboard-3", "Technical", "Consumers exceeding sanctioned load", "count", 15, 380, "count"),
    ("Dashboard-3", "Technical", "% Consumers with Load Violation", "pct", 0.5, 12, "%"),
    ("Dashboard-3", "Technical", "Load Duration Curve & Asset Loading Spread", "pct", 55, 88, "%"),
    ("Dashboard-3", "Technical", "DT Failure Rate (%)", "pct", 0.2, 5.5, "%"),
    ("Dashboard-3", "Technical", "Top Overloaded Assets", "count", 8, 42, "count"),
    ("Dashboard-3", "Technical", "Top Power Quality Issues", "count", 6, 35, "count"),
    ("Dashboard-4", "Operation", "Voltage Deviation (%)", "pct", 1, 12, "%"),
    ("Dashboard-4", "Operation", "Voltage Deviation Index (VDI)", "index", 0.02, 0.95, "index"),
    ("Dashboard-4", "Operation", "Frequency Deviation Index (FDI)", "index", 0.01, 0.35, "index"),
    ("Dashboard-4", "Operation", "Voltage Fluctuation Index", "index", 0.01, 0.45, "index"),
    ("Dashboard-4", "Operation", "Voltage Unbalance Index", "index", 0.02, 0.28, "index"),
    ("Dashboard-4", "Operation", "Voltage Drop (V)", "count", 5, 45, "V"),
    ("Dashboard-4", "Operation", "Low Power Factor (%) by DT/Feeder", "pct", 60, 92, "%"),
    ("Dashboard-4", "Operation", "Meter Current Unbalance (%)", "pct", 2, 18, "%"),
    ("Dashboard-4", "Operation", "% Time beyond voltage tolerance band", "pct", 0.5, 15, "%"),
    ("Dashboard-4", "Operation", "% Time with unacceptable current imbalance (>10%)", "pct", 1, 22, "%"),
    ("Dashboard-5", "Analytics", "Number of Tamper Alerts (Cover Open)", "count", 2, 85, "count"),
    ("Dashboard-5", "Analytics", "Number of Tamper Alerts (External Magnet)", "count", 0, 42, "count"),
    ("Dashboard-5", "Analytics", "Number of Tamper Alerts (Neutral Disturbance)", "count", 1, 38, "count"),
    ("Dashboard-5", "Analytics", "Number of Tamper Alerts (Neutral Missing)", "count", 0, 25, "count"),
    ("Dashboard-5", "Analytics", "Consumption Comparison - Energy Gap (kWh)", "count", 120, 8500, "kWh"),
    ("Dashboard-5", "Analytics", "Total anomalies detected (by time period)", "count", 25, 420, "count"),
    ("Dashboard-5", "Analytics", "Anomalies by type", "count", 3, 12, "types"),
    ("Dashboard-5", "Analytics", "Anomalies by severity", "count", 2, 5, "levels"),
    ("Dashboard-5", "Analytics", "Anomalies by geography", "count", 5, 45, "zones"),
    ("Dashboard-5", "Analytics", "Anomaly trends (daily/weekly/monthly)", "pct", -15, 25, "%"),
    ("Dashboard-5", "Analytics", "Repeat anomaly tracking", "count", 3, 65, "count"),
    ("Dashboard-6", "Analytics", "Theft Suspect Flags", "count", 8, 120, "count"),
    ("Dashboard-6", "Analytics", "% Reduction in Theft Events (monthly trend)", "pct", 5, 45, "%"),
    ("Dashboard-6", "Analytics", "Theft / Load diversion", "count", 2, 55, "count"),
    ("Dashboard-6", "Analytics", "Areas with Highest Theft Risk", "count", 3, 28, "count"),
    ("Dashboard-6", "Finance", "Revenue Recovery Improvement (%)", "pct", 3, 28, "%"),
    ("Dashboard-7", "Analytics", "Communication health issues", "count", 5, 95, "count"),
    ("Dashboard-7", "Technical", "Signal strength statistics", "pct", 72, 98, "%"),
    ("Dashboard-7", "Technical", "Packet loss percentage", "pct", 0.2, 8, "%"),
    ("Dashboard-7", "Technical", "Communication retry counts", "count", 50, 850, "count"),
    ("Dashboard-7", "Technical", "Non-reporting meters (>24 hours)", "count", 12, 220, "count"),
    ("Dashboard-7", "Technical", "Communication technology performance (RF/GPRS/PLC)", "pct", 85, 99, "%"),
    ("Dashboard-7", "Technical", "Weak Signal Percentage", "pct", 2, 18, "%"),
    ("Dashboard-8", "Advanced Analytics", "Auto-indexing consumers and DTRs for correct mapping", "count", 1200, 45000, "count"),
    ("Dashboard-8", "Advanced Analytics", "Track updated tag of DTs to Feeders", "count", 85, 1200, "count"),
    ("Dashboard-8", "Advanced Analytics", "Track updated tag of consumers to DTs", "count", 250, 8500, "count"),
    ("Dashboard-8", "Advanced Analytics", "Re-index consumer/DTR data for correct past-period T&D loss", "count", 500, 12000, "count"),
    ("Dashboard-8", "Advanced Analytics", "Mapping Accuracy (95%)", "pct", 88, 98, "%"),
    ("Dashboard-8", "Advanced Analytics", "DT-to-meter mapping accuracy", "pct", 90, 99, "%"),
    ("Dashboard-8", "Advanced Analytics", "% meters pending field verification (<5%)", "pct", 0.8, 6, "%"),
    ("Dashboard-8", "Advanced Analytics", "Confidence scoring (High/Medium/Low)", "pct", 75, 95, "%"),
    ("Dashboard-8", "Advanced Analytics", "Total assets tracked (Meters/Feeders/DTs)", "count", 5000, 85000, "count"),
    ("Dashboard-8", "Advanced Analytics", "Overloaded DTs identified and monitored", "count", 15, 180, "count"),
    ("Dashboard-8", "Advanced Analytics", "Mismatch analysis (Feeder→DT, DT→Meter)", "count", 20, 450, "count"),
    ("Dashboard-8", "Advanced Analytics", "Correctly mapped meters (%)", "pct", 88, 99, "%"),
    ("Dashboard-8", "Advanced Analytics", "Incorrectly mapped meters requiring correction (%)", "pct", 0.5, 8, "%"),
    ("Dashboard-8", "Advanced Analytics", "Verification pending count", "count", 50, 1200, "count"),
    ("Dashboard-8", "Advanced Analytics", "Correction cycle time (avg days)", "count", 2, 18, "days"),
    ("Dashboard-8", "Advanced Analytics", "Transformer utilization rate (% of rated capacity)", "pct", 45, 88, "%"),
    ("Dashboard-8", "Advanced Analytics", "Field verification completion rate", "pct", 82, 99, "%"),
    ("Dashboard-9", "Analytics", "Tamper sequence detection", "count", 5, 75, "count"),
    ("Dashboard-9", "Analytics", "Voltage/Current imbalance", "count", 8, 95, "count"),
    ("Dashboard-9", "Analytics", "Power factor deterioration", "count", 3, 42, "count"),
    ("Dashboard-9", "Analytics", "Overload / MD breach risk", "count", 12, 88, "count"),
    ("Dashboard-9", "Analytics", "Hidden outage pockets", "count", 2, 35, "count"),
    ("Dashboard-9", "Analytics", "Data quality issues", "count", 15, 120, "count"),
    ("Dashboard-9", "Analytics", "Reverse flow", "count", 0, 28, "count"),
    ("Dashboard-9", "Analytics", "Consumption spikes/drops", "count", 20, 180, "count"),
    ("Dashboard-9", "Analytics", "Phase-level mapping accuracy", "pct", 82, 98, "%"),
    ("Dashboard-9", "Analytics", "Phase imbalance reduced by minimum 30%", "pct", 28, 55, "%"),
    ("Dashboard-9", "Analytics", "Real-time phase load monitoring per transformer", "pct", 85, 99, "%"),
    ("Dashboard-9", "Analytics", "Imbalance alerts when threshold exceeded", "count", 10, 95, "count"),
    ("Dashboard-9", "Analytics", "Phase transfer recommendations (what-if)", "count", 5, 65, "count"),
]


def generate_value(spec):
    _, _, _, vtype, lo, hi, unit = spec
    if vtype == "pct":
        return round(random.uniform(lo, hi), 2)
    if vtype == "count":
        return random.randint(int(lo), int(hi))
    if vtype == "index":
        return round(random.uniform(lo, hi), 3)
    if vtype == "minutes":
        return round(random.uniform(lo, hi), 1)
    if vtype == "kva":
        return random.choice([25, 63, 100, 160, 200, 250, 315])
    if vtype == "score":
        return random.randint(int(lo), int(hi))
    return round(random.uniform(lo, hi), 2)


def autofit_column_widths(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in range(1, min(ws.max_row + 1, 500)):
            cell = ws[f"{letter}{row}"]
            if cell.value:
                val = str(cell.value)
                max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[letter].width = max(max_len + 2, 10)


def main():
    if pd is None:
        raise SystemExit("Install pandas and openpyxl: pip install pandas openpyxl")

    rows = []
    for spec in KPI_SPECS:
        dashboard, dept, kpi_name, _, _, _, unit = spec
        value = generate_value(spec)
        formula, required_data, columns_used, formula_logic = find_formula(kpi_name)
        rows.append({
            "Dashboard Name": dashboard,
            "Department": dept,
            "KPI Name": kpi_name,
            "Exact Formula": formula,
            "Required Data (Profile + Columns)": required_data,
            "Columns Used": columns_used,
            "Formula / Logic Used": formula_logic,
            "Value": value,
            "Unit": unit,
            "Period": "Jan 2025",
        })

    df = pd.DataFrame(rows)
    out_dir = Path(__file__).resolve().parent
    out_path = out_dir / "KPI_Dummy_Dataset.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="KPI_Data", index=False)
        summary = df.groupby("Dashboard Name").agg(
            KPIs=("KPI Name", "count"),
            Departments=("Department", lambda x: x.nunique()),
        ).reset_index()
        summary.to_excel(writer, sheet_name="Summary_by_Dashboard", index=False)

    wb = load_workbook(out_path)
    for sheet_name in wb.sheetnames:
        autofit_column_widths(wb[sheet_name])
    wb.save(out_path)

    print(f"Generated: {out_path}")
    print(f"Total rows: {len(df)}")
    return out_path


if __name__ == "__main__":
    main()
